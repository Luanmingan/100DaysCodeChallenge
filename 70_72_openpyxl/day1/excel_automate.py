from openpyxl import load_workbook
import os


dir_path = os.path.abspath(os.path.dirname(__file__))
xl_filepath = os.path.join(dir_path, 'financial_sample.xlsx')

wb = load_workbook(xl_filepath)
ws1 = wb['Finances 2017']


# This is better than that from the tutorial.
profit_total = 0
col_L = ws1['L']
for cell in col_L[1:101]:
    profit_total += float(cell.value)
print(profit_total)

print('Accessing many cells: ')

print('-'*40)
col_A = ws1['A']
for row in col_A[:10]:
    print(row.value)

print('-'*40)
row_1 = ws1[1]
for col in row_1:
    print(col.value)

print('-'*40)
cell_range = ws1['A1':'B3']
for col in cell_range:
    for row in col:
        print(row)

print('-'*40)
for row in ws1.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
    for cell in row:
        print(cell)

print('-'*40)
for col in ws1.iter_cols(min_row=1, max_row=3, min_col=1, max_col=2):
    for cell in col:
        print(cell)
